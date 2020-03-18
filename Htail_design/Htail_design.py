# -*- coding: utf-8 -*-
"""
Created on Mon Jul  8 08:57:25 2019

@author: 01384599
"""

# In[]
import math
import numpy
import matplotlib.pyplot as plt
import openpyxl
def Exceldivide(file_dir):
    data=openpyxl.load_workbook(file_dir)         #打开excel表
    sheet=data.get_sheet_by_name('Sheet1')
    
    S=sheet.cell(row=9,column=1).value
    c=sheet.cell(row=9,column=2).value
    v=sheet.cell(row=3,column=1).value
    rho=sheet.cell(row=3,column=2).value
    tail_capacity=sheet.cell(row=31,column=1).value
    l_h=sheet.cell(row=31,column=2).value
    x_cm=sheet.cell(row=9,column=3).value
    x_cg_cm=sheet.cell(row=6,column=1).value
    alpha_design_wing=sheet.cell(row=9,column=4).value
    cam=sheet.cell(row=9,column=5).value
    lamb_h=sheet.cell(row=31,column=3).value  #平尾梢根比
    A_h=sheet.cell(row=31,column=4).value   #平尾展弦比
    sweep_angle=sheet.cell(row=31,column=5).value #平尾前缘后掠角
    alpha=[]
    cl_wing=[]
    cm_wing=[]
    cz_wing=[]
    cl_body=[]
    cm_body=[]
    cz_body=[]
    for x in range(9,28):
        if sheet.cell(row=x,column=6).value!=999:
            alpha.append(sheet.cell(row=x,column=6).value)
        if sheet.cell(row=x,column=7).value!=999:
            cl_wing.append(sheet.cell(row=x,column=7).value)
        if sheet.cell(row=x,column=8).value!=999:
            cz_wing.append(sheet.cell(row=x,column=8).value)
        if sheet.cell(row=x,column=9).value!=999:
            cm_wing.append(sheet.cell(row=x,column=9).value)
        if sheet.cell(row=x,column=10).value!=999:
            cl_body.append(sheet.cell(row=x,column=10).value)
        if sheet.cell(row=x,column=11).value!=999:
            cz_body.append(sheet.cell(row=x,column=11).value)
        if sheet.cell(row=x,column=12).value!=999:
            cm_body.append(sheet.cell(row=x,column=12).value)
    return S,c,v,rho,tail_capacity,l_h,x_cm,x_cg_cm,alpha_design_wing,cam,lamb_h,A_h,sweep_angle,alpha,cl_wing,cm_wing,cz_wing,cl_body,cm_body,cz_body
[S,c,v,rho,tail_capacity,l_h,x_cm,x_cg_cm,alpha_design_wing,cam,lamb_h,A_h,sweep_angle,alpha,cl_wing,cm_wing,cz_wing,cl_body,cm_body,cz_body]\
=Exceldivide('F:\\MR项目\\V3.0总体\\气动力设计\\平尾设计程序\\Htail_design.xlsx')
#g.save('F:\\MR项目\\V3.0总体\\气动力设计\\Htail_design.xlsx')  
#from mpl_toolkits.mplot3d import axes3d
#输入
#S=0.625
#c=0.2728
#v=20
#rho=1.225
#tail_capacity=0.8   #平尾容量
#l_h=0.8 #平尾力臂
#x_cm=0   #取矩点x坐标
#x_cg_cm=0    #重心距参考点的X距离，重心在后为正
#alpha_design_wing=4 #机翼设计迎角
#cam=4   #机翼翼型弯度百分值
#alpha=[-4,-2,0,2,4,6,8,10,12,14,16]
#cl_wing=[-0.020472576,0.14162138,0.304123299,0.466007169,0.62568871,0.781612596,
#         0.932174955,1.07438883,1.201180401,1.303005383,1.283660234]
#cl_body=numpy.transpose([0]*11)
#cz_wing=[-0.021562298,0.140964036,0.304123299,0.466570081,0.626394504,0.781725864,
#         0.930668569,1.069996849,1.192532898,1.288996966,1.270932084]
#cz_body=numpy.transpose([0]*11)
#cm_wing=[-0.079391791,-0.134346573,-0.189553999,-0.244488681,-0.298349628,-0.350350151,
#         -0.399724293,-0.445130653,-0.483562775,-0.511936228,-0.515370598]
#cm_body=numpy.transpose([0]*11)
#lamb_h=1/2  #平尾梢根比
#A_h=5    #平尾展弦比
#sweep_angle=5   #平尾前缘后掠角，单位°
#输出
q=0.5*rho*v**2
S_h=tail_capacity*S*c/l_h
b_h=(S_h*A_h)**0.5
c_h=(S_h/A_h)**0.5
root_c_h=c_h*2*(1/lamb_h)/(1/lamb_h+1)
tip_c_h=c_h*2/(1/lamb_h+1)
sweep_angle=sweep_angle/180*math.pi

cl_wing=numpy.array(cl_wing)
cl_body=numpy.array(cl_body)
cz_wing=numpy.array(cz_wing)
cz_body=numpy.array(cz_body)
cm_wing=numpy.array(cm_wing)
cm_body=numpy.array(cm_body)
cm_wing_cg=cm_wing+cz_wing*x_cg_cm/c
cm_body_cg=cm_body+cz_body*x_cg_cm/c
cl=cl_body+cl_wing
cm_cg=cm_body_cg+cm_wing_cg
m_cg=cm_cg*q*S*c
x_ac_cm=(cm_wing[alpha.index(4)]-cm_wing[alpha.index(0)])/(cl_wing[alpha.index(4)]-cl_wing[alpha.index(0)])*c    #机翼焦点距取矩点x距离，焦点在前为正
L_h=m_cg/(l_h-x_ac_cm-x_cg_cm) #平尾配平所需升力
cl_h=L_h/q/S_h
cl_design_h=cl_h[alpha.index(alpha_design_wing)]    #巡航状态下平尾配平所需升力系数

#考虑下洗影响,计算平尾处下洗角
epsilon=35*cl_wing[alpha.index(4)]/(S/c**2)*(1+0.0065*alpha_design_wing)*(1+0.05*cam)   #单位°
cl_alpha_h=2*math.pi/(1+2*(1+2*lamb_h)/A_h/(1+lamb_h))
inc_h=cl_design_h/cl_alpha_h*180/math.pi+epsilon  #尾翼安装角，单位°

#绘制平尾平面形状
shape1=[0,-b_h/2]
shape2=[0,-b_h/2*math.tan(sweep_angle)]
shape3=[-b_h/2,-b_h/2]
shape4=[-b_h/2*math.tan(sweep_angle),-b_h/2*math.tan(sweep_angle)-tip_c_h]
shape5=[-b_h/2,0]
shape6=[-b_h/2*math.tan(sweep_angle)-tip_c_h,-root_c_h]
shape7=[0,b_h/2]
shape8=[0,-b_h/2*math.tan(sweep_angle)]
shape9=[b_h/2,b_h/2]
shape10=[-b_h/2*math.tan(sweep_angle),-b_h/2*math.tan(sweep_angle)-tip_c_h]
shape11=[b_h/2,0]
shape12=[-b_h/2*math.tan(sweep_angle)-tip_c_h,-root_c_h]
plt.plot(shape1,shape2)
plt.plot(shape3,shape4)
plt.plot(shape5,shape6)
plt.plot(shape7,shape8)
plt.plot(shape9,shape10)
plt.plot(shape11,shape12)
plt.xticks(numpy.arange(-b_h/2,b_h/2,b_h/2))
plt.yticks(numpy.arange(-b_h,0,b_h/10))
print('平尾设计升力系数=',cl_design_h)
print('平尾安装角=',inc_h)
#end
